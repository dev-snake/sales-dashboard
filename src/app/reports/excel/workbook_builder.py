"""Excel multi-sheet workbook builder (openpyxl)."""

from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from app.reports.package import ReportDataPackage

HEADER_FILL = PatternFill("solid", fgColor="2563EB")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=14)


def _write_df(ws: Any, df: pd.DataFrame, start_row: int = 1) -> None:
    if df is None or df.empty:
        ws.cell(start_row, 1, "(no data)")
        return
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(r_idx, c_idx, value)
            if r_idx == start_row:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="center")


def _auto_width(ws: Any) -> None:
    for col in ws.columns:
        max_len = 0
        letter = col[0].column_letter
        for cell in col:
            value = cell.value
            max_len = max(max_len, len(str(value if value is not None else "")))
        ws.column_dimensions[letter].width = min(max_len + 2, 40)


def build_excel(package: ReportDataPackage, path: Path) -> Path:
    """Write multi-sheet workbook; return path."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # Summary
    ws = wb.active
    assert ws is not None
    ws.title = "Summary"
    ws["A1"] = package.title
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Generated: {package.generated_at.isoformat()}"
    ws["A3"] = f"Period: {package.period.start} → {package.period.end} (end exclusive)"

    kpi = package.kpi
    prev = package.kpi_previous
    rows = [
        ("Metric", "Current", "Previous", "Delta %"),
        ("Revenue", float(kpi.revenue), float(prev.revenue), package.kpi_delta_pct("revenue")),
        (
            "Gross Profit",
            float(kpi.gross_profit),
            float(prev.gross_profit),
            package.kpi_delta_pct("gross_profit"),
        ),
        (
            "Gross Margin %",
            float(kpi.gross_margin_pct),
            float(prev.gross_margin_pct),
            package.kpi_delta_pct("gross_margin_pct"),
        ),
        ("Orders", kpi.order_count, prev.order_count, package.kpi_delta_pct("order_count")),
        ("Buyers", kpi.buyer_count, prev.buyer_count, package.kpi_delta_pct("buyer_count")),
        ("AOV", float(kpi.aov), float(prev.aov), package.kpi_delta_pct("aov")),
        ("Units", kpi.units_sold, prev.units_sold, package.kpi_delta_pct("units_sold")),
        (
            "Repeat rate %",
            package.repeat_rate.get("repeat_customer_rate_pct"),
            None,
            None,
        ),
    ]
    for r_idx, row in enumerate(rows, start=5):
        for c_idx, val in enumerate(row, start=1):
            cell_value: str | float | int | None = (
                val if val is None or isinstance(val, (str, float, int)) else str(val)
            )
            cell = ws.cell(r_idx, c_idx, cell_value)
            if r_idx == 5:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT

    sheets: list[tuple[str, pd.DataFrame]] = [
        ("Trend", package.trend),
        ("TopProducts", package.top_products),
        ("TopStores", package.top_stores),
        ("TopCustomers", package.top_customers),
        ("TopEmployees", package.top_employees),
        ("Categories", package.by_category),
        ("Regions", package.by_region),
        ("RFM", package.rfm_summary),
        ("ABC", package.abc.head(100) if not package.abc.empty else package.abc),
        (
            "LowStock",
            package.low_stock.head(100) if not package.low_stock.empty else package.low_stock,
        ),
    ]

    if "cohort_matrix" in package.extras:
        sheets.append(("Cohort", package.extras["cohort_matrix"]))

    for name, df in sheets:
        if package.report_type == "daily" and name in {"RFM", "ABC", "Cohort", "Regions"}:
            continue
        ws_s = wb.create_sheet(name)
        _write_df(ws_s, df if isinstance(df, pd.DataFrame) else pd.DataFrame())
        _auto_width(ws_s)

    # Definitions
    defs = wb.create_sheet("Definitions")
    defs["A1"] = "Metric definitions"
    defs["A1"].font = TITLE_FONT
    definitions = [
        ("Revenue", "SUM(order_items.line_total) for paid/completed orders"),
        ("COGS", "SUM(quantity * unit_cost)"),
        ("Gross Profit", "Revenue - COGS"),
        ("Gross Margin %", "Gross Profit / Revenue * 100"),
        ("AOV", "Revenue / order count"),
        ("RFM", "Recency, Frequency, Monetary segments"),
        ("ABC", "Pareto classes A≤80%, B≤95%, C rest of cumulative revenue"),
    ]
    for i, (k, v) in enumerate(definitions, start=3):
        defs.cell(i, 1, k)
        defs.cell(i, 2, v)
    _auto_width(defs)
    _auto_width(ws)

    wb.save(path)
    return path
