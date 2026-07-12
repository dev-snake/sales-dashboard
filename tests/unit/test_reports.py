"""Unit tests for report periods and Excel/PDF builders (no DB)."""

from __future__ import annotations

from datetime import UTC, date, datetime
from decimal import Decimal
from pathlib import Path

import pandas as pd
import pytest

from app.reports.excel.workbook_builder import build_excel
from app.reports.package import ReportDataPackage
from app.reports.pdf.pdf_builder import build_pdf
from app.reports.periods import parse_report_type, resolve_period
from app.schemas.filters import KPIResult


def test_resolve_daily() -> None:
    p = resolve_period("daily", date(2024, 6, 15))
    assert p.start == date(2024, 6, 15)
    assert p.end == date(2024, 6, 16)
    assert p.label == "2024-06-15"
    prev = p.previous
    assert prev.end == p.start
    assert (prev.end - prev.start).days == 1


def test_resolve_monthly() -> None:
    p = resolve_period("monthly", date(2024, 2, 10))
    assert p.start == date(2024, 2, 1)
    assert p.end == date(2024, 3, 1)
    assert p.label == "2024-02"


def test_resolve_quarterly_yearly() -> None:
    q = resolve_period("quarterly", date(2024, 5, 1))
    assert q.start == date(2024, 4, 1)
    assert q.end == date(2024, 7, 1)
    assert q.label == "2024-Q2"
    y = resolve_period("yearly", date(2024, 8, 1))
    assert y.start == date(2024, 1, 1)
    assert y.end == date(2025, 1, 1)


def test_resolve_weekly_iso() -> None:
    # 2024-01-03 is Wednesday → week starts Monday 2024-01-01
    p = resolve_period("weekly", date(2024, 1, 3))
    assert p.start == date(2024, 1, 1)
    assert p.end == date(2024, 1, 8)


def test_parse_report_type() -> None:
    assert parse_report_type("Monthly") == "monthly"
    with pytest.raises(ValueError):
        parse_report_type("hourly")


def _sample_package() -> ReportDataPackage:
    period = resolve_period("monthly", date(2024, 6, 15))
    trend = pd.DataFrame(
        {
            "period_key": pd.date_range("2024-06-01", periods=5, freq="D"),
            "revenue": [100, 120, 90, 140, 130],
        }
    )
    products = pd.DataFrame(
        {
            "sku": ["A", "B", "C"],
            "product_name": ["Alpha", "Beta", "Gamma"],
            "revenue": [500, 300, 200],
            "units": [5, 3, 2],
            "profit": [200, 100, 50],
        }
    )
    stores = pd.DataFrame({"store_code": ["S1", "S2"], "revenue": [800, 200], "orders": [10, 3]})
    return ReportDataPackage(
        report_type="monthly",
        period=period,
        generated_at=datetime(2024, 6, 16, tzinfo=UTC),
        kpi=KPIResult(
            revenue=Decimal("1000"),
            cogs=Decimal("400"),
            gross_profit=Decimal("600"),
            gross_margin_pct=Decimal("60.00"),
            order_count=13,
            buyer_count=8,
            units_sold=10,
            aov=Decimal("76.92"),
        ),
        kpi_previous=KPIResult(
            revenue=Decimal("800"),
            gross_profit=Decimal("450"),
            order_count=10,
            aov=Decimal("80"),
        ),
        trend=trend,
        top_products=products,
        top_stores=stores,
        top_customers=pd.DataFrame(
            {"customer_code": ["C1"], "lifetime_revenue": [400], "order_count": [3]}
        ),
        by_category=pd.DataFrame({"category_name": ["Electronics"], "revenue": [700]}),
        rfm_summary=pd.DataFrame(
            {"segment": ["Champions", "Loyal"], "customers": [2, 5], "monetary": [300, 500]}
        ),
        abc=products.assign(cum_pct=[0.5, 0.8, 1.0], abc_class=["A", "A", "B"]),
        repeat_rate={"total_buyers": 8, "repeat_buyers": 3, "repeat_customer_rate_pct": 37.5},
    )


def test_build_excel(tmp_path: Path) -> None:
    path = build_excel(_sample_package(), tmp_path / "monthly_2024-06.xlsx")
    assert path.is_file()
    assert path.stat().st_size > 1000
    # openpyxl can read back
    from openpyxl import load_workbook

    wb = load_workbook(path)
    assert "Summary" in wb.sheetnames
    assert "TopProducts" in wb.sheetnames
    assert "Definitions" in wb.sheetnames


def test_build_pdf(tmp_path: Path) -> None:
    path = build_pdf(_sample_package(), tmp_path / "monthly_2024-06.pdf")
    assert path.is_file()
    assert path.stat().st_size > 1000
    # PDF magic
    assert path.read_bytes()[:4] == b"%PDF"
